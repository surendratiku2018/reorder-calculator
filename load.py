"""One-time data migration — loads the workbook into SQLite, mirroring the
sheet's structure so a full export reproduces its exact layout.

  * products        — Create Date, Description, Category, Supplier, Lead Time
                      (split: a number -> lead_time; a word like 'Static' ->
                      calc_method, and the row does not calculate).
  * comparison_week — D:J ("LWeek Day 1-7", this year) and K:Q ("LYear Day1-7",
                      last year) stored as explicit per-SKU values.
  * daily_usage     — the dated block (column X onward), at real calendar dates.
  * calc_parameters — safety_factor 1.15 and the detected anchor date.

SKUs are handled as text throughout; no spreadsheet round-trip.
"""

import datetime as _dt

import openpyxl

import db

COL_CREATE_DATE = 1            # A
COL_CATEGORY, COL_SUPPLIER = 2, 3
COL_LWEEK = range(4, 11)       # D..J this year
COL_LYEAR = range(11, 18)      # K..Q last year
COL_SKU, COL_DESC, COL_LEAD = 19, 20, 22
COL_DAILY_START = 24           # X

_MONTHS = {
    month: index + 1
    for index, month in enumerate(
        ["jan", "feb", "mar", "apr", "may", "jun",
         "jul", "aug", "sep", "oct", "nov", "dec"]
    )
}


def _num(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _parse_lead(value):
    """Return ``(lead_time, calc_method)`` for the lead-time cell."""
    if value is None or str(value).strip() == "":
        return None, None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = int(value)
        return (number, None) if number >= 1 else (None, str(value).strip())

    text = str(value).strip()
    try:
        number = int(float(text))
        return (number, None) if number >= 1 else (None, text)
    except ValueError:
        return None, text


def _create_date(value):
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    return None if value is None else str(value).strip()


def _block_dates(labels, start_year=2025):
    """Return ISO dates for the daily-usage headers.

    Supports both real Excel dates/datetimes and legacy labels such as
    ``10-Feb``. For legacy labels, ``start_year`` is used initially and the
    year advances when the month wraps from December to January.
    """
    dates = []
    year = int(start_year)
    previous_month = None

    for header in labels:
        if header is None or str(header).strip() == "":
            dates.append(None)
            continue

        if isinstance(header, (_dt.datetime, _dt.date)):
            value = header.date() if isinstance(header, _dt.datetime) else header
            dates.append(value.isoformat())
            previous_month = value.month
            year = value.year
            continue

        text = str(header).strip().replace(".", "")
        parsed = None

        try:
            parsed = _dt.datetime.fromisoformat(text).date()
        except ValueError:
            try:
                parsed = _dt.date.fromisoformat(text)
            except ValueError:
                parsed = None

        if parsed is not None:
            dates.append(parsed.isoformat())
            previous_month = parsed.month
            year = parsed.year
            continue

        parts = text.split("-")
        try:
            day = int(parts[0])
            month = _MONTHS[parts[1][:3].lower()]
        except (ValueError, KeyError, IndexError):
            dates.append(None)
            continue

        if previous_month is not None and month < previous_month:
            year += 1

        previous_month = month
        dates.append(_dt.date(year, month, day).isoformat())

    return dates


def migrate(xlsx_path, db_path=db.DEFAULT_DB_PATH, sheet="Sheet1", start_year=2025):
    """Load a master spreadsheet into a fresh database.

    Returns an import report containing counts for products, blank SKU rows,
    duplicate SKU rows, suppliers, comparison rows, and daily usage rows.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    conn = None

    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Worksheet '{sheet}' was not found in the workbook.")

        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise ValueError("The uploaded workbook is empty.")

        header = rows[0]
        if len(header) < COL_SKU:
            raise ValueError(
                f"The workbook does not contain the required SKU column "
                f"(column {COL_SKU})."
            )

        block_dates = _block_dates(header[COL_DAILY_START - 1:], start_year)
        anchor_iso = next(
            (date_value for date_value in block_dates if date_value),
            f"{start_year}-01-01",
        )

        report = {
            "products": 0,
            "suppliers": 0,
            "daily_rows": 0,
            "comparison_rows": 0,
            "calc_methods": {},
            "duplicate_skus": [],
            "total_rows": 0,
            "unique_skus": 0,
            "duplicates": 0,
            "blank_rows": 0,
        }

        conn = db.connect(db_path)
        db.reset_db(conn)

        supplier_ids = {}
        seen = set()
        products = []
        comparison_rows = []
        daily_agg = {}

        total_excel_rows = 0
        blank_sku_rows = 0
        duplicate_rows = 0

        for excel_row_number, row in enumerate(rows[1:], start=2):
            total_excel_rows += 1

            sku_value = row[COL_SKU - 1] if len(row) >= COL_SKU else None
            if sku_value is None or str(sku_value).strip() == "":
                blank_sku_rows += 1
                continue

            sku = str(sku_value).strip()

            if sku in seen:
                duplicate_rows += 1
                report["duplicate_skus"].append(
                    {"row": excel_row_number, "sku": sku}
                )
                continue

            seen.add(sku)

            supplier_raw = row[COL_SUPPLIER - 1] if len(row) >= COL_SUPPLIER else None
            supplier_id = None
            if supplier_raw not in (None, "", 0, "0"):
                supplier_name = str(supplier_raw).strip()
                if supplier_name:
                    if supplier_name not in supplier_ids:
                        cursor = conn.execute(
                            "INSERT INTO suppliers(name) VALUES (?)",
                            (supplier_name,),
                        )
                        supplier_ids[supplier_name] = cursor.lastrowid
                    supplier_id = supplier_ids[supplier_name]

            category_value = row[COL_CATEGORY - 1] if len(row) >= COL_CATEGORY else None
            category = (
                str(category_value).strip()
                if category_value is not None
                else None
            )

            lead_value = row[COL_LEAD - 1] if len(row) >= COL_LEAD else None
            lead_time, calc_method = _parse_lead(lead_value)
            report["calc_methods"][calc_method] = (
                report["calc_methods"].get(calc_method, 0) + 1
            )

            description_value = row[COL_DESC - 1] if len(row) >= COL_DESC else None
            create_date_value = (
                row[COL_CREATE_DATE - 1]
                if len(row) >= COL_CREATE_DATE
                else None
            )

            products.append(
                (
                    sku,
                    _create_date(create_date_value),
                    str(description_value).strip()
                    if description_value is not None
                    else None,
                    category,
                    lead_time,
                    calc_method,
                    supplier_id,
                )
            )

            comparison_rows.append(
                (
                    sku,
                    [
                        _num(row[column - 1]) if len(row) >= column else None
                        for column in COL_LWEEK
                    ],
                    [
                        _num(row[column - 1]) if len(row) >= column else None
                        for column in COL_LYEAR
                    ],
                )
            )

            if len(row) >= COL_DAILY_START:
                for index, cell in enumerate(row[COL_DAILY_START - 1:]):
                    quantity = _num(cell)
                    if (
                        quantity not in (None, 0.0)
                        and index < len(block_dates)
                        and block_dates[index]
                    ):
                        key = (sku, block_dates[index])
                        daily_agg[key] = daily_agg.get(key, 0.0) + quantity

        conn.executemany(
            "INSERT INTO products(" 
            "sku, create_date, description, category, lead_time, calc_method, supplier_id"
            ") VALUES (?,?,?,?,?,?,?)",
            products,
        )

        for sku, last_week, last_year in comparison_rows:
            db.set_comparison_week(conn, sku, last_week, last_year)

        conn.executemany(
            "INSERT INTO daily_usage(sku, usage_date, qty) VALUES (?,?,?)",
            [(sku, usage_date, qty) for (sku, usage_date), qty in daily_agg.items()],
        )

        db.set_params(conn, 1.15, anchor_iso)
        conn.commit()

        report["products"] = len(products)
        report["suppliers"] = len(supplier_ids)
        report["daily_rows"] = len(daily_agg)
        report["comparison_rows"] = len(comparison_rows)
        report["total_rows"] = total_excel_rows
        report["unique_skus"] = len(seen)
        report["duplicates"] = duplicate_rows
        report["blank_rows"] = blank_sku_rows

        return report

    except Exception:
        if conn is not None:
            conn.rollback()
        raise
    finally:
        if conn is not None:
            conn.close()
        wb.close()
