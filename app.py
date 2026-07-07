"""Reorder Calculator — Streamlit UI (Technical Spec §7).

Run:  streamlit run app.py

Screens (sidebar):
  • Reorder dashboard — sortable/filterable table of the latest run
  • Run calculation   — safety factor, then recompute with current windows
  • Import usage       — weekly roll-forward (new week -> D:J, front week -> K:Q,
                          anchor advances 7 days), then recompute
  • Edit products      — edit lead time / category / supplier
  • Export             — Excel-safe .xlsx (SKU as text) + full-data export + CSV
"""

import datetime as _dt
import hmac
import os
import tempfile

import pandas as pd
import streamlit as st

import calc
import db
import exports
import rollforward

st.set_page_config(page_title="Reorder Calculator", page_icon="◧", layout="wide")


# ---------- optional password gate ----------
def _expected_password() -> str:
    try:
        pw = st.secrets.get("app_password", "")
    except Exception:
        pw = ""
    return pw or os.environ.get("APP_PASSWORD", "")


def check_password() -> bool:
    expected = _expected_password()
    if not expected:
        return True
    if st.session_state.get("auth_ok"):
        return True

    def _verify():
        entered = st.session_state.get("pw", "")
        st.session_state["auth_ok"] = bool(entered) and hmac.compare_digest(
            entered.encode("utf-8"), expected.encode("utf-8")
        )
        if st.session_state["auth_ok"]:
            st.session_state.pop("pw", None)

    st.title("◧ Reorder Calculator")
    st.text_input("Password", type="password", key="pw", on_change=_verify)
    if st.session_state.get("auth_ok") is False:
        st.error("Incorrect password.")
    st.caption("This internal tool is password-protected. Enter the password to continue.")
    return False


def conn():
    return db.connect(db.DEFAULT_DB_PATH)


def db_ready(c) -> bool:
    try:
        return c.execute("SELECT COUNT(*) FROM products").fetchone()[0] > 0
    except Exception:
        return False


# ---------- data ----------

def dashboard_rows(c, run_date):
    rows = c.execute(
        """
        SELECT p.sku, p.description, COALESCE(s.name,'') AS supplier, p.category,
               p.lead_time, p.calc_method,
               r.baseline_reorder_point, r.yoy_difference, r.new_reorder_point
        FROM products p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        LEFT JOIN reorder_results r ON r.sku = p.sku AND r.run_date = ?
        ORDER BY p.sku
        """,
        (run_date,),
    ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        # one Lead Time column: the number of days, or the word (Static / Sales Velocity)
        d["lead_time_display"] = str(d["lead_time"]) if d["lead_time"] is not None else (d["calc_method"] or "")
        d["calculates"] = d["lead_time"] is not None
        out.append(d)
    return out


# ---------- pages ----------

def page_dashboard(c):
    st.subheader("Reorder dashboard")
    run_date = db.latest_run_date(c)
    if not run_date:
        st.info("No calculation run yet. Go to **Run calculation** and click Run.")
        return
    rows = dashboard_rows(c, run_date)
    df = pd.DataFrame(rows)
    df["sku"] = df["sku"].astype(str)   # keep SKU text — never coerced

    c1, c2, c3, c4 = st.columns([3, 2, 2, 2])
    search = c1.text_input("Search SKU / description", "")
    cats = ["(all)"] + sorted({r["category"] for r in rows if r["category"]})
    sups = ["(all)"] + sorted({r["supplier"] for r in rows if r["supplier"]})
    f_cat = c2.selectbox("Category", cats)
    f_sup = c3.selectbox("Supplier", sups)
    hide_nc = c4.checkbox("Only calculating rows", value=False,
                          help="Hide rows whose Lead Time is a word (Static / Sales Velocity) and so don't calculate")

    view = df.copy()
    if search:
        s = search.lower()
        view = view[view["sku"].str.lower().str.contains(s) | view["description"].fillna("").str.lower().str.contains(s)]
    if f_cat != "(all)":
        view = view[view["category"] == f_cat]
    if f_sup != "(all)":
        view = view[view["supplier"] == f_sup]
    if hide_nc:
        view = view[view["calculates"]]

    display = view[["sku", "description", "supplier", "category", "lead_time_display",
                    "baseline_reorder_point", "yoy_difference", "new_reorder_point"]]
    st.caption(f"Run date **{run_date}** · showing **{len(view)}** of {len(df)} products")
    st.dataframe(
        display, width="stretch", hide_index=True, height=560,
        column_config={
            "sku": st.column_config.TextColumn("Product id"),
            "description": st.column_config.TextColumn("Description", width="large"),
            "supplier": "Supplier",
            "category": "Category",
            "lead_time_display": st.column_config.TextColumn("Lead Time"),
            "baseline_reorder_point": st.column_config.NumberColumn("Baseline (W)"),
            "yoy_difference": st.column_config.NumberColumn("YoY diff (R)", format="%.3g"),
            "new_reorder_point": st.column_config.NumberColumn("New Reorder Point (U)", format="%.3f"),
        },
    )
    st.caption("Rows with a word in Lead Time (Static / Sales Velocity) intentionally show no reorder point.")


def page_run(c):
    st.subheader("Run calculation")
    params = db.get_params(c)
    st.markdown(f"**Baseline anchor (column X):** `{params['lead_window_anchor']}` — the daily block starts here.")
    st.caption("The weekly roll-forward (Import usage) is what advances the anchor and updates the LWeek/LYear columns. Use this screen to recompute with the current data, or to adjust the safety factor.")
    with st.form("run"):
        sf = st.number_input("Safety factor", value=float(params["safety_factor"]), step=0.01, format="%.2f")
        if st.form_submit_button("▶ Run calculation", type="primary"):
            db.set_params(c, sf, params["lead_window_anchor"])
            prev_run = db.latest_run_date(c)
            prev = {r["sku"]: r["new_reorder_point"] for r in
                    c.execute("SELECT sku, new_reorder_point FROM reorder_results WHERE run_date = ?", (prev_run,))} if prev_run else {}
            run_date = _dt.date.today().isoformat()
            n = calc.run_calculation(c, run_date, db.get_params(c))
            now = {r["sku"]: r["new_reorder_point"] for r in
                   c.execute("SELECT sku, new_reorder_point FROM reorder_results WHERE run_date = ?", (run_date,))}
            changed = sum(1 for k, v in now.items() if (v or 0) != (prev.get(k) or 0))
            st.success(f"Computed {n} products for run **{run_date}**. New Reorder Point changed for **{changed}** SKU(s) since the prior run.")


def page_import(c):
    st.subheader("Import usage — weekly roll-forward")
    params = db.get_params(c)
    old_anchor = params["lead_window_anchor"]
    st.markdown(
        "Drop in a week of actual usage. This does exactly the roll-forward you described:\n"
        f"1. the new week's 7 days → **LWeek (D:J)** (this week);\n"
        f"2. the week at the front of the daily block (`{old_anchor}`, last year's same week) → **LYear (K:Q)**;\n"
        f"3. the new week is also **banked onto the end of the daily block** (at its real dates), so it becomes next year's LYear data;\n"
        f"4. the anchor advances 7 days (`{old_anchor}` → `{(_dt.date.fromisoformat(old_anchor)+_dt.timedelta(days=7)).isoformat()}`).\n\n"
        "Prior history is never deleted. SKUs are read strictly as text (stdlib csv — never a spreadsheet engine)."
    )
    st.code("SKU CODE,10-Feb,11-Feb,12-Feb,13-Feb,14-Feb,15-Feb,16-Feb\n36-0023,12,18,24,12,,,\n35-0089,8,12,32,8,8,4,8", language="text")
    up = st.file_uploader("Weekly usage CSV (SKU CODE + 7 day columns)", type=["csv"])
    if up is None:
        return
    text = up.getvalue().decode("utf-8-sig", errors="replace")
    weekly, labels, rep = rollforward.parse_weekly_csv(text)
    if not rep["format_ok"] or not weekly:
        st.error("Couldn't read this as a weekly file. Expect a header row 'SKU CODE' + 7 date columns, then one row per SKU.")
        return
    st.write(f"Parsed **{rep['rows']}** SKU rows · day columns: {', '.join(labels)}")
    nonzero = sum(1 for v in weekly.values() if any(v))
    st.caption(f"{nonzero} SKUs have usage this week; the rest are zero.")
    if st.button("▶ Roll forward & recalculate", type="primary"):
        summary = rollforward.roll_forward(c, weekly, labels)
        st.success(
            f"Rolled forward. New week → **LWeek (D:J)** and banked onto the block end at "
            f"**{summary['new_week_start']} … {summary['new_week_end']}** ({summary['banked_values']} values); "
            f"the front week (`{summary['old_anchor']}`) → **LYear (K:Q)**; anchor advanced "
            f"**{summary['old_anchor']} → {summary['new_anchor']}** (so `{summary['new_anchor']}` is now column X). "
            f"Recomputed {summary['computed']} products."
        )
        if summary["skipped_unknown"]:
            st.warning(f"{len(summary['skipped_unknown'])} SKU(s) in the file aren't in the catalogue and were skipped: {summary['skipped_unknown'][:15]}")


def page_edit(c):
    st.subheader("Edit products")
    st.caption("Lead Time is a number of days (set by the supplier). If it instead holds a word like Static or Sales Velocity, that row does not calculate a reorder point.")
    skus = [r["sku"] for r in c.execute("SELECT sku FROM products ORDER BY sku")]
    sku = st.selectbox("Product id (SKU)", skus)
    p = dict(c.execute("SELECT * FROM products WHERE sku = ?", (sku,)).fetchone())
    suppliers = c.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()
    sup_names = ["(none)"] + [s["name"] for s in suppliers]
    cur_sup = next((s["name"] for s in suppliers if s["id"] == p["supplier_id"]), "(none)")
    current_lead = str(p["lead_time"]) if p["lead_time"] is not None else (p["calc_method"] or "")
    with st.form("edit"):
        desc = st.text_input("Description", p["description"] or "")
        col1, col2 = st.columns(2)
        lead = col1.text_input("Lead Time (number of days — or a word like Static / Sales Velocity)", current_lead)
        cat = col2.text_input("Category", p["category"] or "")
        sup = col1.selectbox("Supplier", sup_names, index=sup_names.index(cur_sup) if cur_sup in sup_names else 0)
        if st.form_submit_button("Save", type="primary"):
            lead_str = lead.strip()
            if lead_str.isdigit():
                lead_val, method = int(lead_str), None
            else:
                lead_val, method = None, (lead_str or None)
            sup_id = next((s["id"] for s in suppliers if s["name"] == sup), None)
            c.execute("UPDATE products SET description=?, lead_time=?, calc_method=?, category=?, supplier_id=? WHERE sku=?",
                      (desc, lead_val, method, cat or None, sup_id, sku))
            c.commit()
            st.success(f"Saved {sku}. Re-run the calculation to update its reorder point.")


def page_export(c):
    st.subheader("Export")
    run_date = db.latest_run_date(c)
    if not run_date:
        st.info("Run a calculation first.")
        return
    st.markdown(
        "**Excel-safe `.xlsx`** is the one to open in Excel/Sheets — the SKU column is typed as Text, "
        "so `01-2500` can never turn into a date. (A CSV opened in a spreadsheet can re-guess types — that's "
        "what produced the dates you saw; use the .xlsx to avoid it.)"
    )

    def xlsx_bytes(builder):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx"); tmp.close()
        builder(c, tmp.name, run_date)
        with open(tmp.name, "rb") as fh:
            return fh.read()

    col1, col2 = st.columns(2)
    col1.download_button("⤓ Results (Excel-safe .xlsx)", xlsx_bytes(exports.results_xlsx),
                         file_name=f"reorder_results_{run_date}.xlsx",
                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    col2.download_button("⤓ Full data incl. all daily columns (.xlsx)", xlsx_bytes(exports.full_data_xlsx),
                         file_name=f"reorder_full_data_{run_date}.xlsx",
                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.divider()
    st.caption("CSV (for tooling/import; may re-coerce if opened directly in a spreadsheet):")
    col3, col4 = st.columns(2)
    col3.download_button("⤓ Results (CSV)", exports.results_csv(c, run_date),
                         file_name=f"reorder_results_{run_date}.csv", mime="text/csv")
    col4.download_button("⤓ Finale subset (CSV)", exports.results_csv(c, run_date, subset=True),
                         file_name=f"finale_upload_{run_date}.csv", mime="text/csv")
    st.caption("Finale subset = Product id + New Reorder Point. Send me your exact Finale column layout and I'll match it.")


def page_startfresh(c):
    st.subheader("Start fresh — load a new master dataset")
    st.warning(
        "⚠️ This **replaces all current data** with a fresh master spreadsheet "
        "(same column layout as your original). Any weekly imports since the last "
        "load are cleared. Use it to re-base the tool on a newer dataset — e.g. "
        "start over with June 2026 data."
    )
    up = st.file_uploader("Master spreadsheet (.xlsx — same layout as your original)", type=["xlsx"])
    if up is None:
        return

    import io
    import openpyxl
    import load
    try:
        wb = openpyxl.load_workbook(io.BytesIO(up.getvalue()), read_only=True, data_only=True)
        sheet = "Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        first_label = header[load.COL_DAILY_START - 1]
        n_products = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                         if r[load.COL_SKU - 1] not in (None, ""))
        wb.close()
    except Exception as e:
        st.error(f"Couldn't read that as a master spreadsheet (expects the original layout): {e}")
        return

    st.write(f"Detected **{n_products}** products. The first daily column (column X) is labelled **{first_label}**.")
    try:
        if isinstance(first_label, (_dt.datetime, _dt.date)):
            first_date = first_label.date() if isinstance(first_label, _dt.datetime) else first_label
            x_day, x_month = first_date.day, first_date.month
            guess_year = first_date.year
        else:
            label_text = str(first_label).strip().replace(".", "")
            try:
                first_date = _dt.datetime.fromisoformat(label_text).date()
                x_day, x_month, guess_year = first_date.day, first_date.month, first_date.year
            except ValueError:
                parts = label_text.split("-")
                x_day, x_month = int(parts[0]), load._MONTHS[parts[1][:3].lower()]
                guess_year = _dt.date.today().year - 1
    except Exception:
        x_day = x_month = None
        guess_year = _dt.date.today().year
    start_year = st.number_input("Year that column X falls in", value=guess_year, step=1, format="%d",
                                 help="The daily headers carry no year. The daily block then spans ~one year from column X.")
    if x_month:
        st.caption(f"→ Column X will be **{_dt.date(int(start_year), x_month, x_day).isoformat()}**, and the block runs ~one year forward from there.")

    if st.checkbox("I understand this overwrites all current data") and st.button("Replace data & rebuild", type="primary"):
        master_path = os.path.join(db.PROJECT_DIR, "data", "source_workbook.xlsx")
        with open(master_path, "wb") as fh:
            fh.write(up.getvalue())
        with st.spinner("Loading the new dataset and recalculating…"):
            c.close()
            rep = load.migrate(master_path, start_year=int(start_year))
            nc = conn()
            calc.run_calculation(nc, _dt.date.today().isoformat())
            anchor = db.get_params(nc)["lead_window_anchor"]
        st.success(f"Started fresh: loaded {rep['products']} products; column X = {anchor}. Recalculated.")
        st.rerun()


# ---------- shell ----------

def main():
    if not check_password():
        return
    c = conn()
    st.sidebar.title("◧ Reorder Calculator")
    if not db_ready(c):
        # Fresh host (e.g. first load on Streamlit Cloud): build the database from
        # the bundled workbook, then run the first calculation.
        xlsx = os.path.join(db.PROJECT_DIR, "data", "source_workbook.xlsx")
        if not os.path.exists(xlsx):
            st.title("Database not built")
            st.code("python load_cli.py data/source_workbook.xlsx", language="bash")
            return
        with st.spinner("Setting up the database from the source workbook (first load only)…"):
            c.close()
            import load
            load.migrate(xlsx)
            c = conn()
            calc.run_calculation(c, _dt.date.today().isoformat())
        st.success("Database built. Loading…")
        st.rerun()

    counts = c.execute("SELECT (SELECT COUNT(*) FROM products), (SELECT COUNT(*) FROM suppliers), (SELECT MAX(run_date) FROM reorder_results)").fetchone()
    st.sidebar.metric("Products", counts[0])
    st.sidebar.metric("Suppliers", counts[1])
    st.sidebar.caption(f"Last run: {counts[2] or '—'}")
    page = st.sidebar.radio("Go to", ["Reorder dashboard", "Run calculation", "Import usage",
                                      "Edit products", "Export", "Start fresh"])

    st.title("Reorder Calculator")
    {"Reorder dashboard": page_dashboard, "Run calculation": page_run,
     "Import usage": page_import, "Edit products": page_edit, "Export": page_export,
     "Start fresh": page_startfresh}[page](c)


main()
