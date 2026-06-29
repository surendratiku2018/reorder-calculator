"""Exports.

Formats:
  • Excel-safe .xlsx — SKU column typed as Text (@) and written as a string, so
    opening in Excel/Sheets can never turn '01-2500' into 'Jan-00'. (xlsx text
    cells are not re-coerced on open — only CSV import re-guesses types.)
  • CSV — for tooling/import; prefer the .xlsx for spreadsheet use.

Scopes:
  • results   — the summary table (attributes + W / R / U).
  • full data — the full sheet layout: Create Date, Category, Supplier,
    LWeek Day 1-7, LYear Day1-7, YoY Difference, Product id, Description,
    New Reorder Point, Lead Time, Reorder point (baseline), then the dated daily
    block starting at column X = the current anchor.
"""

import csv
import io

import openpyxl
from openpyxl.styles import Font, PatternFill

import db

SUMMARY_COLS = [
    ("sku", "Product id"),
    ("description", "Description"),
    ("supplier", "Supplier"),
    ("category", "Category"),
    ("lead_display", "Lead Time"),
    ("baseline_reorder_point", "Baseline (W)"),
    ("yoy_difference", "YoY Difference (R)"),
    ("new_reorder_point", "New Reorder Point (U)"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")


def _summary_rows(conn, run_date):
    rows = conn.execute(
        """
        SELECT p.sku, p.description, COALESCE(s.name,'') AS supplier, p.category,
               p.lead_time, p.calc_method,
               r.baseline_reorder_point, r.yoy_difference, r.new_reorder_point
        FROM products p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        LEFT JOIN reorder_results r ON r.sku = p.sku AND r.run_date = ?
        ORDER BY p.sku
        """,
        (run_date,),
    ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["lead_display"] = str(d["lead_time"]) if d["lead_time"] is not None else (d["calc_method"] or "")
        out.append(d)
    return out


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def results_xlsx(conn, path, run_date=None):
    run_date = run_date or db.latest_run_date(conn)
    rows = _summary_rows(conn, run_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reorder Results"
    ws.append([h for _, h in SUMMARY_COLS])
    for d in rows:
        ws.append([d.get(k) for k, _ in SUMMARY_COLS])
        ws.cell(row=ws.max_row, column=1, value="" if d["sku"] is None else str(d["sku"])).number_format = "@"
    _style_header(ws, len(SUMMARY_COLS))
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    wb.save(path)
    return path


def _fmt_date(iso):
    y, m, d = iso.split("-")
    return f"{int(d):02d}-{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][int(m)-1]}-{y}"


def full_data_xlsx(conn, path, run_date=None):
    """Reproduce the source sheet's full layout, with the daily block starting at
    column X = the current anchor (only 2025/2026 dates)."""
    run_date = run_date or db.latest_run_date(conn)
    anchor = db.get_params(conn)["lead_window_anchor"]

    # daily block = dates at or after the anchor (column X onward)
    dates = [d for (d,) in conn.execute(
        "SELECT DISTINCT usage_date FROM daily_usage WHERE usage_date >= ? ORDER BY usage_date", (anchor,))]
    usage = {}
    for sku, dt, qty in conn.execute("SELECT sku, usage_date, qty FROM daily_usage WHERE usage_date >= ?", (anchor,)):
        usage.setdefault(sku, {})[dt] = qty

    rows = conn.execute(
        """
        SELECT p.sku, p.create_date, p.category, COALESCE(s.name,'') AS supplier,
               p.description, p.lead_time, p.calc_method,
               cw.lweek_d1, cw.lweek_d2, cw.lweek_d3, cw.lweek_d4, cw.lweek_d5, cw.lweek_d6, cw.lweek_d7,
               cw.lyear_d1, cw.lyear_d2, cw.lyear_d3, cw.lyear_d4, cw.lyear_d5, cw.lyear_d6, cw.lyear_d7,
               r.yoy_difference, r.new_reorder_point, r.baseline_reorder_point
        FROM products p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        LEFT JOIN comparison_week cw ON cw.sku = p.sku
        LEFT JOIN reorder_results r ON r.sku = p.sku AND r.run_date = ?
        ORDER BY p.sku
        """,
        (run_date,),
    ).fetchall()

    headers = (["Create Date", "Category", "Supplier"]
               + [f"LWeek Day {i}" for i in range(1, 8)]
               + [f"LYear Day{i}" for i in range(1, 8)]
               + ["YoY Difference", "Product id", "Description", "New Reorder Point",
                  "Lead Time", "Reorder point"]
               + [_fmt_date(d) for d in dates])
    SKU_COL = 19  # Product id position (matches the sheet's column S)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    ws.append(headers)
    for r in rows:
        lead_display = str(r["lead_time"]) if r["lead_time"] is not None else (r["calc_method"] or "")
        line = [r["create_date"], r["category"], r["supplier"]]
        line += [r[f"lweek_d{i}"] for i in range(1, 8)]
        line += [r[f"lyear_d{i}"] for i in range(1, 8)]
        line += [r["yoy_difference"], str(r["sku"]), r["description"], r["new_reorder_point"],
                 lead_display, r["baseline_reorder_point"]]
        per = usage.get(r["sku"], {})
        line += [per.get(d) for d in dates]
        ws.append(line)
        ws.cell(row=ws.max_row, column=SKU_COL).number_format = "@"   # Product id as Text

    _style_header(ws, len(headers))
    ws.freeze_panes = "X2"   # freeze the A-W attribute columns; daily block scrolls
    ws.column_dimensions["S"].width = 16
    ws.column_dimensions["T"].width = 42
    wb.save(path)
    return path


def results_csv(conn, run_date=None, subset=False):
    run_date = run_date or db.latest_run_date(conn)
    rows = _summary_rows(conn, run_date)
    cols = [("sku", "Product id"), ("new_reorder_point", "ReorderPoint")] if subset else SUMMARY_COLS
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_ALL, lineterminator="\n")
    w.writerow([h for _, h in cols])
    for d in rows:
        out = []
        for k, _ in cols:
            v = d.get(k)
            if isinstance(v, float) and v == int(v):
                v = int(v)
            out.append("" if v is None else v)
        w.writerow(out)
    return buf.getvalue()
